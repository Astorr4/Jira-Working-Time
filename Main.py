import os
import time
import urllib3
import keyring
import getpass
import calendar
import win32com.client
from jira import JIRA
import dateutil.parser
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
# Отлючаем варнинги из-за TLS сертификата
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
# Формируем переменные с датами за прошлый месяц
# Текущая дата
today = date.today()
# Если текущий месяц январь, то переключаемся на декабрь прошлого года
if today.month == 1:
    year = today.year - 1
    month = 12
else:
    year = today.year
    month = today.month - 1  # Если выгружаем за текущий месяц необходимо убрать "- 1"
# Первый день прошлого месяца
first_day = date(year, month, 1)
# Последний день прошлого месяца
last_day = date(year, month, calendar.monthrange(year, month)[1])
print(first_day, last_day)
# Название файла excel в виде прошлого месяца
file_path = f'{calendar.month_name[month]}.xlsx'
# Список наших сотрудников
people = ['Список', 'ФИО', 'Сотрудников']
# Берём наши креды
login = getpass.getuser()
password = keyring.get_password("Jira", login)
# Наш jql запрос
jql = f'(project in (prokect1, prokect2) OR (PROJECT in (prokect3) AND Level = name)) AND created >= {first_day} AND created <= {last_day} ORDER BY created DESC'
# Подключаемся
jira = JIRA('https://jiradomain.ru/jira',
            basic_auth=(login, password), options={'verify': False})
search_result = jira.search_issues(jql, maxResults=10000)
# Счетчик задач
i = 0
# Общее кол-во задач в месяце
l = len(search_result)


def write_dict_to_excel(data_dict):
    # Проверяем существование файла
    file_exists = os.path.exists(file_path)
    if file_exists:
        # Загружаем существующий файл
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # Создаем новый файл
        wb = Workbook()
        ws = wb.active
        # Пишем заголовки (ключи словаря)
        headers = list(data_dict.keys())
        ws.append(headers)
    # Пишем данные
    try:
        ws.append(list(data_dict.values()))
    except Exception as e:
        print(f"Ошибка записи данных в Excel: {e}")
    # Сохраняем файл
    wb.save(file_path)


def date_time_format(datetime):
    try:
        return dateutil.parser.isoparse(datetime).strftime("%d.%m.%Y %H:%M")
    except:
        return None


def send_email():
    # Создание экземпляра приложения Outlook
    outlook = win32com.client.Dispatch('Outlook.Application')
    # Создание нового сообщения
    mail = outlook.CreateItem(0)
    mail.Subject = f'Выгрузка трудозатрат Jira за {month}.{year}'
    # Основной текст письма с использованием HTML для форматирования
    body_text = f"""
            Файл с выгрузкой трудозатрат Jira за прошлый месяц во вложении
        """
    mail.HTMLBody = body_text
    # Добавление основных адресатов
    mail.To = "; ".join(
        ["user@domain.ru"])
    # Указание адреса отправителя
    mail.SentOnBehalfOfName = "user@domain.ru"  # Замените на нужный адрес
    # Добавление вложения
    mail.Attachments.Add(fr'C:\file\path\{file_path}')
    # Отправка письма
    mail.Send()


def write_task_worklog_to_excel(issue):
    # Создаём пустой словарь
    worklog_dict = {}
    # Заполняем словарь общими для трудосписаний данными
    worklog_dict["Ключ"] = issue.key
    worklog_dict["Название"] = issue.fields.summary
    worklog_dict["Проект"] = "АС «АС»" if issue.fields.project.name == "SUPPORT" else "АС «ААСС»"
    worklog_dict["Дата и время обращения"] = date_time_format(
        getattr(issue.fields, 'customfield_13501', None))
    worklog_dict["Дата и время создания"] = date_time_format(
        issue.fields.created)
    worklog_dict["Дата и время оформления"] = date_time_format(issue.fields.customfield_13600) if worklog_dict[
        "Проект"] == "АС «АС»" else None
    worklog_dict["Дата и время резолюции"] = date_time_format(
        issue.fields.resolutiondate)
    worklog_dict["Статус"] = issue.fields.status.name
    worklog_dict["Резолюция"] = getattr(
        getattr(issue.fields, 'resolution', None), 'name', None)
    worklog_dict["Заявитель"] = getattr(
        getattr(issue.fields, 'customfield_10711', None), 'displayName', None)
    # Выделяем ворклоги
    worklog = issue.fields.worklog.worklogs
    # Итерируемся по ворклогам
    for i in range(len(worklog)):
        # Записываем автора и время трудосписания
        worklog_dict["Автор трудосписания"] = worklog[i].author.displayName
        worklog_dict["Дата и время трудосписания"] = dateutil.parser.isoparse(worklog[i].started).strftime(
            "%d.%m.%Y %H:%M")
        worklog_dict["Затрачено (сек)"] = worklog[i].timeSpentSeconds
        if (not hasattr(worklog[i], 'comment') or len(worklog[i].comment) < 2) and worklog_dict[
                "Проект"] == 'АС «АС»' and i == len(worklog) - 1:
            # В АК при закрытии задачи комметарий идёт не в журнал работ, а в отдельное поле "Результат"
            worklog_dict["Комментарий к трудосписанию"] = issue.fields.customfield_11806
        else:
            worklog_dict["Комментарий к трудосписанию"] = getattr(
                worklog[i], 'comment', None)
        if worklog_dict['Автор трудосписания'].strip() in people:
            write_dict_to_excel(worklog_dict)


for issue in search_result:
    issue = jira.issue(issue.key)
    print('{}/{}  {}: {}'.format(i, l, issue.key, issue.fields.summary))
    write_task_worklog_to_excel(issue)
    i += 1
    time.sleep(1)
# Загружаем существующий файл
wb = load_workbook(f"{calendar.month_name[month]}.xlsx")
ws = wb.active
# Автоматическое определение диапазона данных
max_row = ws.max_row
max_col = ws.max_column
table_range = f"A1:{get_column_letter(max_col)}{max_row}"
# Создаем таблицу
table = Table(displayName="AutoTable", ref=table_range)
# Применяем стиль "Средний 9"
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True,
)
table.tableStyleInfo = style
# Добавляем таблицу на лист
ws.add_table(table)
# Устанавливаем стандартную ширину для всех колонок
for col_num in range(1, ws.max_column + 1):  # Проходим по всем колонкам
    # Преобразуем номер колонки в букву
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = 45
# Сохраняем изменения
wb.save(f"{file_path}")
send_email()
